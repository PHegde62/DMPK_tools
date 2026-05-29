"""
reader.py
Reads the CRO Excel workbook tab by tab, guided entirely by column_map.yaml.
Returns a dict of {tab_key: DataFrame} with canonical column names applied.
No assay-specific logic — all tab names, header rows, and column renames
come from the config.
"""

from __future__ import annotations
import re
import pandas as pd
from pathlib import Path
from typing import Any
from openpyxl import load_workbook


# ── Public entry point ────────────────────────────────────────────────────────

def read_cro_workbook(
    workbook_path: str | Path,
    column_map: dict[str, Any],
) -> dict[str, Any]:
    """
    Read all tabs defined in column_map["tabs"] from the CRO workbook.

    Returns
    -------
    dict with keys matching tab keys in column_map (e.g. "summary", "materials")
    Values are DataFrames for tabular tabs, dicts for key_value tabs,
    and raw row lists for compound_blocks tabs (handled by tab_merger).
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CRO workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    available_sheets = {s.lower(): s for s in wb.sheetnames}

    results: dict[str, Any] = {}

    for tab_key, tab_cfg in column_map["tabs"].items():
        sheet_name = tab_cfg["sheet_name"]
        actual_name = available_sheets.get(sheet_name.lower())
        if actual_name is None:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook. "
                f"Available sheets: {wb.sheetnames}"
            )

        layout = tab_cfg.get("layout", "tabular")

        if layout == "key_value":
            results[tab_key] = _read_key_value(wb[actual_name], tab_cfg)
        elif layout == "tabular":
            results[tab_key] = _read_tabular(wb[actual_name], tab_cfg)
        elif layout == "compound_blocks":
            # Raw rows — passed to tab_merger for block parsing
            results[tab_key] = _read_raw_rows(wb[actual_name])
        else:
            raise ValueError(f"Unknown layout '{layout}' for tab '{tab_key}'")

    # Also grab Study Design tab as plain text for text_contains QC checks
    study_design_name = available_sheets.get("study design")
    if study_design_name:
        results["study_design_text"] = _read_full_text(wb[study_design_name])

    wb.close()
    return results


# ── Layout readers ────────────────────────────────────────────────────────────

def _read_key_value(ws, tab_cfg: dict) -> dict[str, Any]:
    """
    Read a key→value sheet (e.g. Signature tab).
    Col A = key, Col B = value. Returns a flat dict with canonical keys.
    """
    raw: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None and row[1] is not None:
            raw[str(row[0]).strip()] = row[1]

    result: dict[str, Any] = {}
    for field in tab_cfg.get("fields", []):
        raw_key = field["raw_key"]
        canonical = field["canonical"]
        result[canonical] = raw.get(raw_key)

    return result


def _read_tabular(ws, tab_cfg: dict) -> pd.DataFrame:
    """
    Read a standard header+data sheet.
    header_row (0-indexed) tells us which row has column names.
    """
    header_row_idx = tab_cfg.get("header_row", 0)
    columns_cfg = tab_cfg.get("columns", [])

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows or header_row_idx >= len(all_rows):
        return pd.DataFrame()

    headers = [str(c).strip() if c is not None else f"_col{i}"
               for i, c in enumerate(all_rows[header_row_idx])]
    data_rows = all_rows[header_row_idx + 1:]

    df = pd.DataFrame(data_rows, columns=headers)

    # Keep and rename only the mapped columns
    rename_map: dict[str, str] = {}
    keep_cols: list[str] = []
    for col_cfg in columns_cfg:
        raw_name = col_cfg["raw"]
        canonical = col_cfg["canonical"]
        # Fuzzy header match — strip µ/u variants and extra spaces
        matched = _fuzzy_find_column(raw_name, headers)
        if matched is None:
            if col_cfg.get("required", False):
                raise ValueError(
                    f"Required column '{raw_name}' not found in sheet "
                    f"'{tab_cfg['sheet_name']}'. Available: {headers}"
                )
            continue
        rename_map[matched] = canonical
        keep_cols.append(matched)

    df = df[keep_cols].rename(columns=rename_map)

    # Apply dtypes
    for col_cfg in columns_cfg:
        if "dtype" in col_cfg:
            canonical = col_cfg["canonical"]
            if canonical in df.columns:
                df[canonical] = pd.to_numeric(df[canonical], errors="coerce")

    # Drop fully-empty rows
    df = df.dropna(how="all").reset_index(drop=True)

    # Drop rows where compound_id looks like a footnote (very long strings)
    if "compound_id" in df.columns:
        df = df[df["compound_id"].astype(str).str.len() <= 50].reset_index(drop=True)

    return df


def _read_raw_rows(ws) -> list[tuple]:
    """Return all rows as a list of value tuples (for compound_blocks parser)."""
    return [row for row in ws.iter_rows(values_only=True)]


def _read_full_text(ws) -> str:
    """Concatenate all non-empty cell values into a single text blob."""
    parts = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                parts.append(str(cell))
    return " ".join(parts)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fuzzy_find_column(raw_name: str, headers: list[str]) -> str | None:
    """
    Match a column name tolerantly:
    1. Exact match
    2. Case-insensitive match
    3. Unicode-normalised match (µ → u, strip spaces)
    """
    # Normalise for comparison
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", s.strip().lower()).replace("μ", "u").replace("µ", "u")

    target = norm(raw_name)
    for h in headers:
        if norm(h) == target:
            return h
    return None
