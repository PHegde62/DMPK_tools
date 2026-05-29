"""
qc_reporter.py
Writes a color-coded Excel QC report from QCResult objects.
Layout mirrors the PC-018 Review Checklist structure (section grouping,
pass/fail/warn/manual color coding, manual-review checkboxes).
"""

from __future__ import annotations
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from modules.rule_evaluator import QCResult
from modules.qc_engine import qc_summary


# ── Colour palette (matches checklist legend) ─────────────────────────────────
FILL = {
    "pass":   PatternFill("solid", fgColor="C6EFCE"),   # green
    "warn":   PatternFill("solid", fgColor="FFEB9C"),   # amber
    "fail":   PatternFill("solid", fgColor="FFC7CE"),   # red
    "manual": PatternFill("solid", fgColor="DDEBF7"),   # blue (review)
    "header": PatternFill("solid", fgColor="2F4F8F"),
    "section":PatternFill("solid", fgColor="BDD7EE"),
}
FONT_WHITE  = Font(bold=True, color="FFFFFF")
FONT_BOLD   = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

STATUS_LABEL = {
    "pass":   "PASS",
    "fail":   "FAIL",
    "warn":   "WARN",
    "manual": "MANUAL REVIEW",
}


def write_qc_report(
    results: list[QCResult],
    output_path: str | Path,
    report_metadata: dict[str, Any] | None = None,
) -> Path:
    """
    Write a QC report Excel file.

    Parameters
    ----------
    results         : list of QCResult from qc_engine.run_qc()
    output_path     : where to write the .xlsx
    report_metadata : optional dict with keys like report_number, cro_name, etc.

    Returns
    -------
    Path of the written file
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Report"

    row = 1

    # ── Title block ───────────────────────────────────────────────────────────
    meta = report_metadata or {}
    _write_cell(ws, row, 1, "DMPK Assay QC Report", fill=FILL["header"],
                font=FONT_WHITE, span=(1, 5))
    row += 1
    for label, key in [
        ("Report Number", "report_number"),
        ("CRO",           "cro_name"),
        ("Study Date",    "study_start_date"),
        ("Assay Type",    "assay_type"),
    ]:
        _write_cell(ws, row, 1, label, font=FONT_BOLD)
        _write_cell(ws, row, 2, meta.get(key, ""))
        row += 1
    row += 1

    # ── Summary block ─────────────────────────────────────────────────────────
    summary = qc_summary(results)
    _write_cell(ws, row, 1, "QC Summary", fill=FILL["header"],
                font=FONT_WHITE, span=(1, 5))
    row += 1
    for label, val in [
        ("Total rules", summary["total"]),
        ("Passed",      summary["passed"]),
        ("Warnings",    summary["warned"]),
        ("Failures",    summary["failed"]),
        ("Manual review required", summary["manual"]),
        ("Upload blocked?", "YES — resolve failures first" if summary["upload_blocked"] else "No"),
    ]:
        _write_cell(ws, row, 1, label, font=FONT_BOLD)
        cell = ws.cell(row=row, column=2, value=val)
        if label == "Upload blocked?" and summary["upload_blocked"]:
            cell.fill = FILL["fail"]
            cell.font = FONT_BOLD
        row += 1
    row += 1

    # ── Rule results table ────────────────────────────────────────────────────
    headers = ["Rule ID", "Section", "Description", "Status", "Detail", "Affected Compounds"]
    for col, h in enumerate(headers, 1):
        _write_cell(ws, row, col, h, fill=FILL["header"], font=FONT_WHITE)
    row += 1

    current_section = None
    for result in results:
        # Section divider row
        if result.section != current_section:
            current_section = result.section
            _write_cell(ws, row, 1, current_section, fill=FILL["section"],
                        font=FONT_BOLD, span=(1, 6))
            row += 1

        fill = FILL.get(result.status, FILL["warn"])

        status_label = STATUS_LABEL.get(result.status, result.status.upper())
        # Manual check: add a checkbox-style prompt
        detail = result.detail
        if result.status == "manual":
            detail = f"☐  {result.detail}"

        values = [
            result.rule_id,
            result.section,
            result.description,
            status_label,
            detail,
            ", ".join(result.affected_compounds) if result.affected_compounds else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        row += 1

    # ── Legend ────────────────────────────────────────────────────────────────
    row += 1
    _write_cell(ws, row, 1, "Legend", font=FONT_BOLD)
    row += 1
    for label, fill_key in [("PASS — acceptable", "pass"), ("WARN — borderline", "warn"),
                              ("FAIL — must resolve before upload", "fail"),
                              ("MANUAL REVIEW — scientist sign-off required", "manual")]:
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = FILL[fill_key]
        cell.border = THIN_BORDER
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 16, 55, 16, 65, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A12"  # freeze title + summary block

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ── Helpers ───────────────────────────────────────────────────────────────────

def _write_cell(ws, row, col, value, fill=None, font=None, span=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.border = THIN_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if span:
        start_col, end_col = span
        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col
        )
