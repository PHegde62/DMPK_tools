"""
upload_writer.py
Maps canonical DataFrame columns → upload template column names (from std_rules.yaml),
enforces the exact column order defined in upload_columns_ordered,
then writes the final upload Excel file.
"""

from __future__ import annotations
import shutil
import pandas as pd
from pathlib import Path
from typing import Any
from openpyxl import load_workbook


def build_upload_df(df: pd.DataFrame, std_rules: dict[str, Any]) -> pd.DataFrame:
    """
    Map canonical column names → upload template column names.
    Columns are returned in the order defined by upload_columns_ordered.
    Missing columns are included as empty (None) so the template structure
    is always complete even when a field wasn't present in the raw data.
    """
    col_map: dict[str, str] = std_rules.get("upload_column_map", {})
    ordered: list[str] = std_rules.get("upload_columns_ordered", list(col_map.values()))

    upload_df = pd.DataFrame()
    for canonical, upload_col in col_map.items():
        if canonical in df.columns:
            upload_df[upload_col] = df[canonical].values
        else:
            upload_df[upload_col] = None

    # Reorder to match upload_columns_ordered, add any missing cols as empty
    final_cols = []
    for col in ordered:
        if col not in upload_df.columns:
            upload_df[col] = None
        final_cols.append(col)

    # Keep only ordered columns, in order
    return upload_df[final_cols]


def write_upload_file(
    upload_df: pd.DataFrame,
    template_path: str | Path,
    output_path: str | Path,
    assay_sheet_name: str | None = None,
) -> Path:
    """
    Write the upload DataFrame into a copy of the upload template Excel file.
    Uses the correct sheet tab for the assay type when the template has
    multiple sheets (one per assay).

    Parameters
    ----------
    upload_df        : DataFrame from build_upload_df()
    template_path    : path to the upload template .xlsx
    output_path      : where to write the populated file
    assay_sheet_name : sheet name in the template to write into.
                       If None, uses the first (active) sheet.
    """
    output_path = Path(output_path)
    template_path = Path(template_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)

    # Select the right sheet
    if assay_sheet_name and assay_sheet_name in wb.sheetnames:
        ws = wb[assay_sheet_name]
    else:
        ws = wb.active

    # Read header row from template
    header = [cell.value for cell in ws[1]]
    col_index = {name: i + 1 for i, name in enumerate(header) if name is not None}

    # Clear existing data rows (keep header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write rows using direct column name lookup (no namedtuple mangling)
    for row_idx, (_, data_row) in enumerate(upload_df.iterrows(), start=2):
        for col_name, col_num in col_index.items():
            val = data_row.get(col_name)
            # Convert NaN/None cleanly
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            ws.cell(row=row_idx, column=col_num, value=val)

    wb.save(output_path)
    return output_path


# Map assay_type → sheet name in the combined upload template
ASSAY_TO_SHEET: dict[str, str] = {
    "microsomal_stability":  "Mic stability",
    "kinetic_solubility":    "Solubility",
    "permeability":          "Permeability",
    "ppb":                   "PPB",
    "logd":                  "LogD",
    "hepatocyte_stability":  "Hep stability",
    "hep_binding":           "Sheet7",
}
